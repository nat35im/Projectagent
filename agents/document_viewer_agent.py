"""
Agent Mesh – Document Viewer Agent
Safely reads uploaded project documents (PDF, DOCX, XLSX) from the
sandboxed data/docs/projects/ directory. Read-only; no path traversal.
"""
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from orchestrator.state import AgentState
from orchestrator.llm_factory import get_llm
from langchain_core.messages import SystemMessage, HumanMessage

llm = get_llm()

AGENT_NAME = "Document Viewer Agent"

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
DOCS_ROOT = os.path.join(PROJECT_ROOT, "data", "docs", "projects")


def _safe_path(project_code: str, filename: str) -> str | None:
    """
    Resolve an absolute path inside the sandboxed DOCS_ROOT.
    Returns None if the resolved path escapes the sandbox (path traversal guard).
    """
    safe_code = os.path.normpath(project_code)
    candidate = os.path.realpath(os.path.join(DOCS_ROOT, safe_code, filename))
    if not candidate.startswith(os.path.realpath(DOCS_ROOT) + os.sep):
        return None
    return candidate


def _list_project_docs(project_code: str) -> list[str]:
    """Return filenames available for a project code."""
    folder = os.path.join(DOCS_ROOT, project_code)
    if not os.path.isdir(folder):
        return []
    return [
        f for f in os.listdir(folder)
        if os.path.isfile(os.path.join(folder, f))
        and not f.startswith(".")
        and f.rsplit(".", 1)[-1].lower() in {"pdf", "docx", "doc", "xlsx", "xls"}
    ]


def _extract_text(filepath: str) -> str:
    """Extract plain text from a PDF, DOCX, or XLSX file (read-only)."""
    ext = filepath.rsplit(".", 1)[-1].lower()

    if ext in {"docx", "doc"}:
        try:
            from docx import Document
            doc = Document(filepath)
            return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        except ImportError:
            return "⚠️ python-docx not installed. Cannot read DOCX files."
        except Exception as e:
            return f"⚠️ Could not read DOCX: {e}"

    if ext == "pdf":
        try:
            import pdfplumber
            with pdfplumber.open(filepath) as pdf:
                pages = [page.extract_text() or "" for page in pdf.pages]
            return "\n\n".join(p for p in pages if p.strip())
        except ImportError:
            pass
        # Fallback: pypdf
        try:
            from pypdf import PdfReader
            reader = PdfReader(filepath)
            return "\n\n".join(
                page.extract_text() or "" for page in reader.pages
            )
        except ImportError:
            return "⚠️ pdfplumber or pypdf not installed. Cannot read PDF files."
        except Exception as e:
            return f"⚠️ Could not read PDF: {e}"

    if ext in {"xlsx", "xls"}:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            parts = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    if any(c.strip() for c in cells):
                        rows.append("\t".join(cells))
                if rows:
                    parts.append(f"### Sheet: {sheet}\n" + "\n".join(rows))
            wb.close()
            return "\n\n".join(parts)
        except ImportError:
            return "⚠️ openpyxl not installed. Cannot read XLSX files."
        except Exception as e:
            return f"⚠️ Could not read XLSX: {e}"

    return f"⚠️ Unsupported file type: .{ext}"


def _infer_project_code_from_query(query: str, history: list[dict]) -> str | None:
    """
    Ask the LLM to extract a project number or customer name from the query,
    then match it against the available project folders.
    """
    if not os.path.isdir(DOCS_ROOT):
        return None

    available_codes = [
        d for d in os.listdir(DOCS_ROOT)
        if os.path.isdir(os.path.join(DOCS_ROOT, d))
    ]
    if not available_codes:
        return None

    combined_q = (query + " " + " ".join(m["content"] for m in history)).lower()
    for code in available_codes:
        if code.lower() in combined_q:
            return code
    return None


def document_viewer_agent_node(state: AgentState) -> dict:
    query = state["query"]
    history = state.get("history", [])
    current_outputs = state.get("agent_outputs", [])
    debug = state.get("debug_log", "")

    # 1. Locate project folder
    project_code = _infer_project_code_from_query(query, history)

    if not project_code:
        # List all available projects with docs
        if os.path.isdir(DOCS_ROOT):
            projects_with_docs = [
                d for d in os.listdir(DOCS_ROOT)
                if os.path.isdir(os.path.join(DOCS_ROOT, d)) and _list_project_docs(d)
            ]
        else:
            projects_with_docs = []

        if projects_with_docs:
            listing = "\n".join(f"- `{p}`" for p in projects_with_docs)
            msg = (
                f"--- {AGENT_NAME} ---\n"
                "I can show documents for these projects:\n"
                f"{listing}\n\n"
                "Please specify a project code or customer name."
            )
        else:
            msg = f"--- {AGENT_NAME} ---\nNo project documents have been uploaded yet."

        return {
            "response": msg,
            "agent_outputs": current_outputs + [msg],
            "debug_log": debug + f"\n🗂️ {AGENT_NAME}: no project code identified.",
        }

    # 2. List documents for this project
    docs = _list_project_docs(project_code)
    if not docs:
        msg = f"--- {AGENT_NAME} ---\nNo documents found for project `{project_code}`."
        return {
            "response": msg,
            "agent_outputs": current_outputs + [msg],
            "debug_log": debug + f"\n🗂️ {AGENT_NAME}: no files in {project_code}.",
        }

    # 3. Determine which file to view
    q_lower = query.lower()
    target_file = None
    for doc in docs:
        if doc.lower() in q_lower or os.path.splitext(doc)[0].lower() in q_lower:
            target_file = doc
            break

    # If query mentions a type, pick the first matching type
    if not target_file:
        for doc in docs:
            ext = doc.rsplit(".", 1)[-1].lower()
            if ("contract" in q_lower or "sow" in q_lower) and ext in {"pdf", "docx", "doc"}:
                target_file = doc
                break
            if ("estimat" in q_lower or "milestone" in q_lower or "excel" in q_lower) and ext in {"xlsx", "xls"}:
                target_file = doc
                break

    # Default: show all available docs listed; read first if only one
    if not target_file:
        if len(docs) == 1:
            target_file = docs[0]
        else:
            listing = "\n".join(f"- `{d}`" for d in docs)
            msg = (
                f"--- {AGENT_NAME} ---\n"
                f"Project `{project_code}` has the following documents:\n{listing}\n\n"
                "Which document would you like to view?"
            )
            return {
                "response": msg,
                "agent_outputs": current_outputs + [msg],
                "debug_log": debug + f"\n🗂️ {AGENT_NAME}: multiple docs, awaiting selection.",
            }

    # 4. Resolve path safely (sandbox guard)
    safe_fp = _safe_path(project_code, target_file)
    if not safe_fp or not os.path.isfile(safe_fp):
        msg = f"--- {AGENT_NAME} ---\n❌ Document `{target_file}` not accessible."
        return {
            "response": msg,
            "agent_outputs": current_outputs + [msg],
            "debug_log": debug + f"\n❌ {AGENT_NAME}: path guard blocked or file missing.",
        }

    # 5. Extract text
    raw_text = _extract_text(safe_fp)
    if not raw_text.strip():
        msg = f"--- {AGENT_NAME} ---\n⚠️ `{target_file}` appears to be empty or unreadable."
        return {
            "response": msg,
            "agent_outputs": current_outputs + [msg],
            "debug_log": debug + f"\n⚠️ {AGENT_NAME}: empty extraction.",
        }

    # 6. Summarise / answer with LLM (keep text within context limits)
    MAX_CHARS = 12_000
    truncated = raw_text[:MAX_CHARS]
    if len(raw_text) > MAX_CHARS:
        truncated += "\n\n*(Document truncated for display — showing first 12 000 characters)*"

    system_prompt = (
        f"You are the {AGENT_NAME}. The user wants to view or ask questions about a project document.\n"
        "Present the document content in a clean, well-structured markdown format.\n"
        "If the user has a specific question about the document, answer it using ONLY the content below.\n"
        "Never invent or add information not present in the document.\n\n"
        f"Document: `{target_file}` (project: {project_code})\n\n"
        f"---\n{truncated}\n---"
    )

    messages = [SystemMessage(content=system_prompt)]
    for msg in history:
        if msg["role"] == "user":
            messages.append(HumanMessage(content=msg["content"]))
        else:
            from langchain_core.messages import AIMessage
            messages.append(AIMessage(content=msg["content"]))
    messages.append(HumanMessage(content=query))

    try:
        response = llm.invoke(messages)
        answer = response.content.strip()
    except Exception as e:
        answer = f"⚠️ Could not summarise document: {e}\n\nRaw content:\n{truncated}"

    report = f"--- {AGENT_NAME}: `{target_file}` ---\n{answer}\n"
    return {
        "response": answer,
        "agent_outputs": current_outputs + [report],
        "debug_log": debug + f"\n📄 {AGENT_NAME}: served `{target_file}` from project `{project_code}`.",
    }
